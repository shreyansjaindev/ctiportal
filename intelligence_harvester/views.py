from rest_framework import filters, viewsets
from django_filters.rest_framework import DjangoFilterBackend
from .models import *
from .serializers import *
from .filters import *
from rest_framework.response import Response
from rest_framework import status
from django.http import FileResponse
import openpyxl
import io
from scripts.identifier import get_indicator_type
from scripts.engine import collect_data, generate_excel
from rest_framework.renderers import BaseRenderer, BrowsableAPIRenderer, JSONRenderer


# Company
class SourceViewSet(viewsets.ModelViewSet):
    queryset = Source.objects.all()
    serializer_class = SourceSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = SourceFilter


class IdentifierViewSet(viewsets.ViewSet):
    serializer_class = IndicatorSerializer

    def create(self, request, *args, **kwargs):
        indicators = {"indicators": request.data.get("data", [])}
        serializer = self.serializer_class(data=indicators)
        if serializer.is_valid():
            indicators = serializer.validated_data["indicators"]
            context = {"query_list": get_indicator_type(indicators)}
            return Response(context, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Constants
EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXCEL_FILENAME = 'attachment; filename="Intelligence_Harvester_Export.xlsx"'


class ExcelRenderer(BaseRenderer):
    media_type = EXCEL_CONTENT_TYPE
    format = "xlsx"
    charset = None
    render_style = "binary"

    def render(self, data, media_type=None, renderer_context=None):
        wb = openpyxl.Workbook()
        wb = generate_excel(wb, data)

        if not wb.sheetnames:
            wb.create_sheet("No Results")

        # Save workbook to a BytesIO object
        output = io.BytesIO()
        wb.save(output)

        output.seek(0)

        # Create a FileResponse with the correct content type and headers
        response = FileResponse(output, content_type=self.media_type)
        response["Content-Disposition"] = "attachment; filename=download.xlsx"
        return response


class SearchViewSet(viewsets.ViewSet):
    serializer_class = SearchSerializer
    renderer_classes = [BrowsableAPIRenderer, JSONRenderer, ExcelRenderer]

    def create(self, request):
        indicators = {"indicators": request.data.get("data", [])}
        serializer = self.serializer_class(data=indicators)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        indicators = serializer.validated_data["indicators"]

        indicator_data_list = [
            {
                "id": indicator.get("id"),
                "value": indicator.get("value"),
                "value_type": indicator.get("type"),
                "sources": indicator.get("checklist"),
            }
            for indicator in indicators
        ]

        results = collect_data(indicator_data_list)

        # results = [
        #     {
        #         "id": indicator.get("id"),
        #         "indicator_data": data,
        #     }
        #     for indicator, data in zip(indicators, collect_data(indicator_data_list))
        # ]

        # results = []

        # for indicator in indicators:
        #     indicator_data = {
        #         "value": indicator.get("value"),
        #         "type": indicator.get("type"),
        #         "sources": indicator.get("checklist"),
        #     }

        #     result = {
        #         "id": indicator.get("id"),
        #         "indicator_data": collect_data(indicator_data),
        #     }

        #     results.append(result)

        return Response({"data": results})
