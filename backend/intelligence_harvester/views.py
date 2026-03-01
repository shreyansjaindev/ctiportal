import openpyxl
import io
import logging
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.renderers import BaseRenderer, JSONRenderer

from api.response import error, success
from scripts.utils.excel_export import generate_excel
from scripts.utils.identifier import get_indicator_type

from .serializers import BatchIndicatorLookupSerializer, IndicatorSerializer
from .services.lookups import execute_batch_lookups
from .services.providers import build_providers_payload


logger = logging.getLogger(__name__)

class IdentifierViewSet(viewsets.ViewSet):
    serializer_class = IndicatorSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        try:
            indicators = {"indicators": request.data.get("indicators", [])}
            serializer = self.serializer_class(data=indicators)
            if serializer.is_valid():
                indicators = serializer.validated_data["indicators"]
                logger.info(f"Identifying {len(indicators)} indicators")
                result = get_indicator_type(indicators)
                logger.debug(f"Identifier result: {result}")
                return success(result)
            else:
                logger.warning(f"Invalid indicators payload: {serializer.errors}")
                return error(
                    "Invalid indicators payload",
                    code="validation_error",
                    details=serializer.errors,
                    status_code=400,
                )
        except Exception as e:
            logger.error(f"Error in identifier: {e}", exc_info=True)
            return error(
                "Failed to identify indicators",
                code="processing_error",
                status_code=500,
            )


class ExcelRenderer(BaseRenderer):
    """Render lookup results as Excel spreadsheet"""
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    format = "xlsx"
    charset = None
    render_style = "binary"

    def render(self, data, media_type=None, renderer_context=None):
        wb = openpyxl.Workbook()
        wb = generate_excel(wb, data)

        if not wb.sheetnames:
            wb.create_sheet("No Results")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Set Content-Disposition via the response in renderer_context.
        # Renderers must return bytes — returning a Response object here is wrong.
        if renderer_context:
            resp = renderer_context.get("response")
            if resp:
                resp["Content-Disposition"] = 'attachment; filename="Intelligence_Harvester_Export.xlsx"'

        return output.read()


class IndicatorLookupViewSet(viewsets.ViewSet):
    """Perform batch indicator lookups across multiple types and providers"""
    serializer_class = BatchIndicatorLookupSerializer
    renderer_classes = [JSONRenderer, ExcelRenderer]
    permission_classes = [IsAuthenticated]

    def create(self, request):
        """
        Execute indicator lookups with specified providers.
        
        Expected request format:
        {
            "indicators": ["example.com"],
            "providers_by_type": {
                "whois": ["whoisxmlapi"],
                "reputation": ["abuseipdb"]
            }
        }
        """
        try:
            serializer = BatchIndicatorLookupSerializer(data=request.data)
            if not serializer.is_valid():
                logger.warning(f"Invalid batch lookup payload: {serializer.errors}")
                return error(
                    "Invalid batch lookup payload",
                    code="validation_error",
                    details=serializer.errors,
                    status_code=400,
                )

            indicators = serializer.validated_data["indicators"]
            providers_by_type = serializer.validated_data["providers_by_type"]
            logger.info(f"Batch lookup request: {len(indicators)} indicators")
            
            # Validate providers_by_type is not empty
            if not providers_by_type:
                logger.warning("No lookup types requested")
                return error(
                    "No lookup types requested",
                    code="validation_error",
                    status_code=400,
                )
            
            lookup_types = list(providers_by_type.keys())
            logger.debug(f"Requested lookup types: {lookup_types}")

            payload = execute_batch_lookups(indicators, providers_by_type)

            logger.info(f"Batch lookup completed: {len(payload['results'])} indicators processed")
            return success(payload)
        except Exception as e:
            logger.error(f"Error in indicator lookup: {e}", exc_info=True)
            return error(
                "Failed to process lookup",
                code="processing_error",
                status_code=500,
            )


class AllProvidersView(viewsets.ViewSet):
    """Get all providers organized by type - unified endpoint"""
    permission_classes = [IsAuthenticated]
    
    def list(self, request):
        try:
            return success(build_providers_payload())
        except Exception as e:
            logger.error(f"Failed to build providers response: {e}")
            return error(f"Failed to load providers: {str(e)}", status_code=500)
